from __future__ import annotations

import csv
import io
import json
import os
import secrets
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from functools import wraps
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, TypeVar

from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from werkzeug.security import check_password_hash, generate_password_hash

from loan_math import FALLBACK_RATES, SOURCE_URL, calculate_statement


BASE_DIR = Path(__file__).resolve().parent
SETTINGS_FILE = BASE_DIR / "instance" / "settings.json"


def load_settings() -> dict[str, Any] | None:
    if not SETTINGS_FILE.exists():
        return None
    return json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))


def save_settings(username: str, display_name: str, password: str) -> None:
    SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "username": username.strip().lower(),
        "display_name": display_name.strip(),
        "password_hash": generate_password_hash(password),
        "secret_key": secrets.token_hex(32),
    }
    SETTINGS_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")


settings = load_settings()
app = Flask(__name__)
app.config.update(
    SECRET_KEY=(settings or {}).get("secret_key", secrets.token_hex(32)),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
)

F = TypeVar("F", bound=Callable[..., Any])


def login_required(view: F) -> F:
    @wraps(view)
    def wrapped(*args: Any, **kwargs: Any) -> Any:
        if not session.get("signed_in"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped  # type: ignore[return-value]


def current_rates() -> tuple[list[dict[str, Any]], bool]:
    try:
        query = urllib.parse.urlencode(
            {
                "csv.x": "yes",
                "Datefrom": "01/Jan/2023",
                "Dateto": "now",
                "SeriesCodes": "IUDBEDR",
                "CSVF": "TN",
                "UsingCodes": "Y",
                "VPD": "Y",
                "VFD": "N",
            }
        )
        url = f"https://www.bankofengland.co.uk/boeapps/database/_iadb-fromshowcolumns.asp?{query}"
        req = urllib.request.Request(url, headers={"User-Agent": "LoanStatementCalculator/1.0"})
        with urllib.request.urlopen(req, timeout=5) as response:
            content = response.read().decode("utf-8-sig")
        rates: list[dict[str, Any]] = []
        previous: float | None = None
        for row in csv.DictReader(io.StringIO(content)):
            rate = float(row["IUDBEDR"])
            if rate != previous:
                effective = datetime.strptime(row["DATE"], "%d %b %Y").date().isoformat()
                rates.append({"date": effective, "rate": rate})
                previous = rate
        if rates:
            return rates, True
    except (KeyError, OSError, TimeoutError, ValueError):
        pass
    return [dict(item) for item in FALLBACK_RATES], False


@app.get("/")
def index() -> Any:
    if load_settings() is None:
        return redirect(url_for("setup"))
    return redirect(url_for("calculator") if session.get("signed_in") else url_for("login"))


@app.route("/setup", methods=["GET", "POST"])
def setup() -> Any:
    global settings
    if load_settings() is not None:
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        display_name = request.form.get("display_name", "").strip()
        password = request.form.get("password", "")
        if not username or not display_name:
            error = "Enter a display name and username."
        elif len(password) < 10:
            error = "Use a password containing at least 10 characters."
        else:
            save_settings(username, display_name, password)
            settings = load_settings()
            app.secret_key = settings["secret_key"]
            session["signed_in"] = True
            return redirect(url_for("calculator"))
    return render_template("setup.html", error=error)


@app.route("/login", methods=["GET", "POST"])
def login() -> Any:
    global settings
    settings = load_settings()
    if settings is None:
        return redirect(url_for("setup"))
    if session.get("signed_in"):
        return redirect(url_for("calculator"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        if username == settings["username"] and check_password_hash(settings["password_hash"], password):
            session.clear()
            session["signed_in"] = True
            return redirect(url_for("calculator"))
        error = "The username or password was not recognised."
    return render_template("login.html", error=error)


@app.post("/logout")
def logout() -> Any:
    session.clear()
    return redirect(url_for("login"))


@app.get("/calculator")
@login_required
def calculator() -> Any:
    settings_now = load_settings() or {}
    return render_template("calculator.html", display_name=settings_now.get("display_name", "Loan owner"))


@app.get("/api/bank-rate")
@login_required
def bank_rate() -> Any:
    rates, live = current_rates()
    return jsonify(
        rates=rates,
        source=SOURCE_URL,
        live=live,
        retrieved_at=datetime.now(timezone.utc).isoformat(),
    )


def statement_from_request() -> tuple[dict[str, Any], list[dict[str, Any]]]:
    data = request.get_json(force=True)
    inputs = {
        "principal": float(data["principal"]),
        "discount": float(data["discount"]),
        "start": str(data["start"]),
        "end": str(data["end"]),
        "timing": str(data.get("timing", "arrears")),
        "rates": data.get("rates") or FALLBACK_RATES,
        "payments": data.get("payments") or [],
    }
    rows = calculate_statement(**inputs)
    return inputs, rows


@app.post("/api/calculate")
@login_required
def calculate() -> Any:
    try:
        _, rows = statement_from_request()
    except (KeyError, TypeError, ValueError) as exc:
        return jsonify(error=str(exc)), 400
    return jsonify(rows=rows)


@app.post("/export")
@login_required
def export() -> Any:
    try:
        inputs, rows = statement_from_request()
    except (KeyError, TypeError, ValueError) as exc:
        return jsonify(error=str(exc)), 400

    workbook = Workbook()
    assumptions = workbook.active
    assumptions.title = "Assumptions"
    assumptions.append(["LOAN STATEMENT"])
    assumptions["A1"].font = Font(bold=True, size=16)
    for item in [
        ("Principal", inputs["principal"]),
        ("Bank Rate source", SOURCE_URL),
        ("Discount", inputs["discount"] / 100),
        ("Start date", inputs["start"]),
        ("End date", inputs["end"]),
        ("Interest timing", inputs["timing"]),
        ("Convention", "Actual/365; daily weighted Bank Rate; monthly anniversary periods"),
        ("No CSV", "Assume each month's calculated interest was paid"),
        ("Missing CSV month", "Repeat the previous month's payment"),
        ("Waterfall", "Accrued interest first, then principal"),
    ]:
        assumptions.append(item)
    assumptions.column_dimensions["A"].width = 24
    assumptions.column_dimensions["B"].width = 74

    statement = workbook.create_sheet("Monthly Statement")
    if rows:
        statement.append(list(rows[0].keys()))
        for row in rows:
            statement.append(list(row.values()))
        for cell in statement[1]:
            cell.font = Font(bold=True)
        for index in range(1, statement.max_column + 1):
            statement.column_dimensions[get_column_letter(index)].width = 18

    payments = workbook.create_sheet("Payments")
    payments.append(["date", "amount"])
    for payment in inputs["payments"]:
        payments.append([payment.get("date", ""), payment.get("amount", "")])

    rates = workbook.create_sheet("Bank Rate History")
    rates.append(["date", "rate", "source"])
    for rate in inputs["rates"]:
        rates.append([rate["date"], float(rate["rate"]) / 100, SOURCE_URL])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"loan-statement-{inputs['start']}-to-{inputs['end']}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(host=os.environ.get("LOAN_HOST", "127.0.0.1"), port=int(os.environ.get("LOAN_PORT", "5000")), debug=False)
